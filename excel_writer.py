import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

SHEET_NAME = "raw_reels"

HEADERS = [
    "scraped_at",
    "discovery_rank",
    "reel_url",
    "reel_id",
    "username",
    "posted_at",
    "caption",
    "audio_name",
    "view_count",
    "like_count",
    "comment_count",
    "source_surface",
    "scrape_status",
]

WRAP_COLUMNS = {"caption"}


def _write_header(sheet: Worksheet):
    for col_index, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.font = Font(bold=True)
        if header in WRAP_COLUMNS:
            sheet.column_dimensions[cell.column_letter].width = 40
        else:
            sheet.column_dimensions[cell.column_letter].width = 18

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(row=1, column=len(HEADERS)).column_letter}1"


def create_workbook_if_missing(path: str):
    if os.path.exists(path):
        return

    os.makedirs(os.path.dirname(path), exist_ok=True)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    sheet = workbook.create_sheet(SHEET_NAME)
    _write_header(sheet)

    workbook.save(path)


def append_row(path: str, row: dict):
    create_workbook_if_missing(path)

    workbook = load_workbook(path)
    sheet = workbook[SHEET_NAME]

    next_row = sheet.max_row + 1
    for col_index, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=next_row, column=col_index, value=row.get(header))
        if header in WRAP_COLUMNS:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(path)


def load_existing_reel_urls(path: str) -> set:
    """Read reel_url values already recorded in the workbook, so a new run
    can seed its dedup set and only append newly observed Reels.
    """
    if not os.path.exists(path):
        return set()

    workbook = load_workbook(path, read_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        return set()

    sheet = workbook[SHEET_NAME]
    reel_url_index = HEADERS.index("reel_url")

    urls = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) > reel_url_index and row[reel_url_index]:
            urls.add(row[reel_url_index])

    return urls


def load_max_discovery_rank(path: str) -> int:
    """Read existing discovery_rank values so a new run continues
    numbering from the max rather than restarting at 1.
    """
    if not os.path.exists(path):
        return 0

    workbook = load_workbook(path, read_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        return 0

    sheet = workbook[SHEET_NAME]
    rank_index = HEADERS.index("discovery_rank")

    max_rank = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) > rank_index and isinstance(row[rank_index], (int, float)):
            max_rank = max(max_rank, int(row[rank_index]))

    return max_rank


def open_workbook_for_update(path: str):
    """Open the workbook read-write for a batch of targeted row updates
    (e.g. a whole enrichment pass) — caller saves once when done.
    """
    workbook = load_workbook(path)
    sheet = workbook[SHEET_NAME]
    return workbook, sheet


def find_row_by_reel_url(sheet: Worksheet, reel_url: str):
    """Return the 1-indexed row number whose reel_url cell matches, or None."""
    reel_url_index = HEADERS.index("reel_url")
    for row_cells in sheet.iter_rows(min_row=2):
        if row_cells[reel_url_index].value == reel_url:
            return row_cells[0].row
    return None


def apply_enrichment_updates(sheet: Worksheet, row_num: int, updates: dict):
    """Fill only currently-blank cells in the row from updates — never
    overwrites an existing non-blank value. If a previously-blank caption
    is now filled and the row's scrape_status is "partial", upgrade it to
    "success" (never the reverse: a "success" row is never downgraded).
    """
    caption_index = HEADERS.index("caption") + 1
    status_index = HEADERS.index("scrape_status") + 1

    caption_was_blank = sheet.cell(row=row_num, column=caption_index).value in (None, "")

    for header, value in updates.items():
        if header not in HEADERS or header == "scrape_status":
            continue
        if value in (None, ""):
            continue

        col_index = HEADERS.index(header) + 1
        cell = sheet.cell(row=row_num, column=col_index)
        if cell.value not in (None, ""):
            continue

        cell.value = value
        if header in WRAP_COLUMNS:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    if caption_was_blank and updates.get("caption"):
        status_cell = sheet.cell(row=row_num, column=status_index)
        if status_cell.value == "partial":
            status_cell.value = "success"
